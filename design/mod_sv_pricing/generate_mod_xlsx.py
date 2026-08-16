#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用 SV 定价生成器：为任意模组生成 Excel 与 meks 预设 JSON。

用法：
    python generate_mod_xlsx.py --name "Twilight Forest" --modid twilightforest \
        --data ../twilightforest_sv/analysis.json --price ../twilightforest_sv/price.json \
        --xlsx ../../docs/TwilightForest_SV定价表.xlsx \
        --preset ../../src/main/resources/data/meks/sv/twilightforest_preset.json
"""
import argparse
import json
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PRESET_EMC_PATH = os.path.join(ROOT, "src", "main", "resources", "data", "meks", "sv", "emc_preset.json")

TAG_FALLBACK = {
    "minecraft:planks": "minecraft:oak_planks",
    "c:planks": "minecraft:oak_planks",
    "c:paper": "minecraft:paper",
    "c:dusts/glowstone": "minecraft:glowstone_dust",
    "c:dusts/redstone": "minecraft:redstone",
    "c:nuggets/gold": "minecraft:gold_nugget",
    "c:ingots/iron": "minecraft:iron_ingot",
    "c:ingots/gold": "minecraft:gold_ingot",
    "c:ingots/copper": "minecraft:copper_ingot",
    "c:rods/wooden": "minecraft:stick",
    "c:rods/blaze": "minecraft:blaze_rod",
    "c:leather": "minecraft:leather",
    "c:string": "minecraft:string",
    "twilightforest:fiery_vial": "twilightforest:fiery_blood",
    "twilightforest:arctic_fur": "twilightforest:arctic_fur",
    "c:ingots/ironwood": "twilightforest:ironwood_ingot",
    "c:ingots/steeleaf": "twilightforest:steeleaf_ingot",
    "c:ingots/knightmetal": "twilightforest:knightmetal_ingot",
    "c:ingots/fiery": "twilightforest:fiery_ingot",
    "c:ingots/silver": "iceandfire:silver_ingot",
    "c:ingots/copper": "minecraft:copper_ingot",
    "c:gems/sapphire": "iceandfire:sapphire_gem",
    "c:bones/wither": "iceandfire:witherbone",
    "iceandfire:scales/dragon/fire": "iceandfire:dragonscales",
    "iceandfire:scales/dragon/ice": "iceandfire:dragonscales",
    "iceandfire:scales/dragon/lightning": "iceandfire:dragonscales",
    "iceandfire:dragon_skulls": "iceandfire:dragon_skull_fire",
}


def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def base_values(price):
    values = load_json(PRESET_EMC_PATH)
    values = {e["item"]: e["emc"] for e in values}
    values["minecraft:end_stone_brick_slab"] = 2
    values["minecraft:totem_of_undying"] = 9216
    values["minecraft:elytra"] = 98304
    for item_id, data in price.items():
        if data.get("sv"):
            values[item_id] = data["sv"]
    return values


def resolve_key(ing, values):
    key = ing.get("item")
    if key:
        return key
    tag = ing.get("tag")
    return TAG_FALLBACK.get(tag) or tag


def derive(values, recipes):
    recipes = [r for r in recipes if r["type"] in ("minecraft:crafting_shaped", "minecraft:crafting_shapeless")]
    derived = {}
    for _ in range(80):
        changed = False
        for r in recipes:
            rid = r["result_id"]
            if rid in values:
                continue
            total = 0
            parts = []
            ok = True
            for ing in r["ingredients"]:
                key = resolve_key(ing, values)
                v = values.get(key)
                if v is None:
                    ok = False
                    break
                total += v * ing.get("count", 1)
                parts.append((key, ing.get("count", 1), v))
            if not ok:
                continue
            per = total // max(1, r["result_count"])
            if per > 0:
                values[rid] = per
                derived[rid] = {
                    "total": total,
                    "count": r["result_count"],
                    "per": per,
                    "parts": parts,
                    "file": r["file"],
                }
                changed = True
        if not changed:
            break
    return derived


def formula_for_derived(info):
    parts = " + ".join(f"{c}×{k}({v})" for k, c, v in info["parts"])
    return f"合成：({parts}) ÷ {info['count']} = {info['per']}"


def source_text(item_id, recipes, loot):
    name = item_id.split(":", 1)[1]
    bits = []
    type_names = {
        "minecraft:crafting_shaped": "合成",
        "minecraft:crafting_shapeless": "合成",
        "minecraft:stonecutting": "切石",
        "minecraft:smithing_transform": "锻造",
        "minecraft:smelting": "烧炼",
        "minecraft:blasting": "高炉",
        "minecraft:smoking": "烟熏",
        "minecraft:campfire_cooking": "营火",
        "twilightforest:drying": "晾晒",
        "twilightforest:scepter_repair": "权杖修复",
        "twilightforest:uncrafting": "拆解台",
        "iceandfire:dragonforge": "龙锻",
    }
    seen_types = set()
    for r in recipes:
        if r["result_id"] == item_id:
            seen_types.add(r["type"])
    for t in sorted(seen_types):
        bits.append(type_names.get(t, t))
    loot_bits = []
    labels = {"chests": "箱子", "entities": "生物掉落", "blocks": "方块掉落", "items": "物品", "archaeology": "考古"}
    for l in loot:
        if item_id in l["items"]:
            parts = l["file"].split("/")
            if len(parts) >= 2:
                kind = labels.get(parts[0], parts[0])
                loot_bits.append(f"{kind}:{parts[-1].removesuffix('.json')}")
    if loot_bits:
        bits.extend(loot_bits)
    return "；".join(bits) if bits else "未找到配方/掉落"


def classify(item, price):
    item_id = item["id"]
    if item_id in price:
        return price[item_id]["cat"]
    name = item["name"]
    if item.get("is_block"):
        if name.endswith("spawner") or "miniature" in name or "trophy_pedestal" in name:
            return "功能方块"
        return "方块"
    if name.endswith("spawn_egg"):
        return "刷怪蛋"
    if name.endswith("bucket"):
        return "桶"
    if "music_disc" in name:
        return "唱片"
    if name.endswith(("_helmet", "_chestplate", "_leggings", "_boots", "_pauldrons")) or "helm" in name:
        return "护甲"
    if any(k in name for k in ("_sword", "_pickaxe", "_axe", "_shovel", "_hoe", "_bow", "_scepter", "_staff",
                                "_flute", "_horn", "_gauntlet", "_claws", "_hammer", "_dagger", "_macuahuitl",
                                "_trident", "_spear", "_targe", "_shield", "_fan", "_magnet", "_meter", "_watch",
                                "_dial", "_bomb", "_seeker", "_stick", "_wand", "_key")):
        return "武器/工具"
    if name.endswith("_trophy"):
        return "奖杯"
    if name.endswith("_banner_pattern"):
        return "旗帜图案"
    if name.endswith("_skull") or name.endswith("_head"):
        return "战利品"
    if name.endswith("_egg") or name.endswith("_egg_giant"):
        return "生物蛋"
    return "材料/其他"


def handling(item_id, price, derived):
    if item_id in price and price[item_id].get("sv"):
        return "显式预设"
    if item_id in derived:
        return "自动推导"
    return "暂不定价"


def build_rows(analysis, price, values, derived):
    rows = []
    recipes = analysis["recipes"]
    loot = analysis["loot"]
    for item in analysis["items"]:
        item_id = item["id"]
        zh = item.get("zh") or item.get("en") or item["name"]
        en = item.get("en") or item["name"]
        src = source_text(item_id, recipes, loot)
        derived_info = derived.get(item_id)
        price_info = price.get(item_id)
        treat = handling(item_id, price, derived)
        if price_info and price_info.get("sv"):
            sv = price_info["sv"]
            logic = price_info["logic"]
            note = price_info.get("note", "")
        elif derived_info:
            sv = None
            logic = formula_for_derived(derived_info)
            note = "运行时按配方自动推导，无需写入预设表。"
        else:
            sv = None
            note = ""
            if price_info:
                logic = price_info["logic"]
                note = price_info.get("note", "")
            elif classify(item, price) == "刷怪蛋":
                logic = "创造模式刷怪蛋，正常生存不可获取，不建议定价。"
            elif classify(item, price) == "桶":
                logic = "实体桶含 NBT，SV 按物品 ID 记录会失真，不建议定价。"
            elif "spawner" in item["name"] or "miniature" in item["name"] or "creative" in item["name"] or "debug" in item["name"]:
                logic = "创造/结构专属物品，不建议定价。"
            else:
                logic = "暂无稳定获取方式或纯装饰/结构物品，暂不定价。"
        rows.append({
            "cat": classify(item, price),
            "id": item_id,
            "zh": zh,
            "en": en,
            "source": src,
            "treat": treat,
            "sv": sv,
            "derived_sv": derived_info["per"] if derived_info else None,
            "logic": logic,
            "note": note,
        })
    return rows


def style_sheet(ws, headers, widths, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--name", required=True)
    parser.add_argument("--modid", required=True)
    parser.add_argument("--data", required=True)
    parser.add_argument("--price", required=True)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--preset", required=True)
    args = parser.parse_args()

    analysis = load_json(args.data)
    price = load_json(args.price)
    values = base_values(price)
    derived = derive(values, analysis["recipes"])
    rows = build_rows(analysis, price, values, derived)

    os.makedirs(os.path.dirname(args.xlsx), exist_ok=True)
    os.makedirs(os.path.dirname(args.preset), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "定价原则"
    overview = [
        ["SV 定价原则", ""],
        ["数据基准", f"{args.name}（1.21.1）；SV 采用 ProjectE EMC 风格，与 emc_preset.json 同一尺度。"],
        ["核心规则", "SV = 获取成本 + 稀有度 + 可重复性；可重复合成品的 SV 不得超过原料成本，否则会形成“合成造币”。"],
        ["自动推导", "原版合成(有型/无序)会被 MeksValues 自动按 Σ(原料SV×数量)÷输出数 推导，无需写入预设表。"],
        ["需显式预设", "锻造、龙锻、晾晒、拆解台等自定义/非原版合成不会自动推导，表格会给出建议值；熔炉/切石装饰多数暂不定价。"],
        ["暂不定价", "刷怪蛋、生物桶、创造专属功能物品、WIP/无稳定获取物品、纯装饰结构物品。"],
        ["", ""],
        ["价格锚点（ProjectE EMC）", ""],
        ["石头", "1"], ["铁锭", "256"], ["钻石", "8192"], ["下界合金锭", "50152"],
        ["下界合金胸甲", "123185"], ["下界之星", "139264"],
    ]
    for row in overview:
        ws.append(row)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("物品定价")
    headers = ["分类", "物品ID", "中文名", "英文名", "获取方式", "建议处理", "建议SV", "自动推导SV", "定价逻辑", "备注"]
    widths = [13, 42, 26, 32, 46, 14, 12, 13, 80, 40]
    ws2.append(headers)
    for r in sorted(rows, key=lambda x: (x["cat"], x["id"])):
        ws2.append([
            r["cat"], r["id"], r["zh"], r["en"], r["source"], r["treat"],
            r["sv"] if r["sv"] else "", r["derived_sv"] if r["derived_sv"] else "",
            r["logic"], r["note"],
        ])
    style_sheet(ws2, headers, widths)

    ws3 = wb.create_sheet("掉落与配方速查")
    headers3 = ["物品ID", "中文名", "掉落来源", "配方类型"]
    widths3 = [42, 26, 60, 44]
    ws3.append(headers3)
    recipes_by_id = defaultdict(list)
    for r in analysis["recipes"]:
        recipes_by_id[r["result_id"]].append(r["type"])
    loot_by_id = defaultdict(list)
    for l in analysis["loot"]:
        for it in l["items"]:
            loot_by_id[it].append(l["file"])
    for item_id in sorted(x for x in set(list(recipes_by_id) + list(loot_by_id)) if x):
        zh = ""
        for item in analysis["items"]:
            if item["id"] == item_id:
                zh = item.get("zh") or item.get("en") or ""
                break
        ws3.append([item_id, zh, "；".join(loot_by_id.get(item_id, [])), "；".join(sorted(set(recipes_by_id.get(item_id, []))))])
    style_sheet(ws3, headers3, widths3)

    wb.save(args.xlsx)

    preset = []
    seen = set()
    for r in rows:
        value = r["sv"] if r["sv"] else r["derived_sv"]
        if not value or r["id"] in seen:
            continue
        seen.add(r["id"])
        preset.append({"item": r["id"], "emc": value})
    preset.sort(key=lambda e: (e["item"].split(":")[0], e["item"]))
    with open(args.preset, "w", encoding="utf-8") as f:
        json.dump(preset, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print("mod", args.modid, "rows", len(rows))
    print("explicit", sum(1 for r in rows if r["sv"]))
    print("derived", sum(1 for r in rows if r["derived_sv"] and not r["sv"]))
    print("unpriced", sum(1 for r in rows if not r["sv"] and not r["derived_sv"]))
    print("preset entries", len(preset))
    print("saved", args.xlsx)
    print("saved", args.preset)


if __name__ == "__main__":
    main()
