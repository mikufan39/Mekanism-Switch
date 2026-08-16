#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 Cataclysm SV 定价表 Excel。

数据基准：
- L_Ender's Cataclysm 1.21.1-3.32（Modrinth jar + GitHub 源码）
- 本模组现有的 ProjectE EMC 预设表 src/main/resources/data/meks/sv/emc_preset.json

用法：
    python generate_cataclysm_sv_xlsx.py
输出：
    docs/Cataclysm_SV定价表.xlsx
"""
import json
import os
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR = os.path.join(ROOT, "design", "cataclysm_sv")
OUT = os.path.join(ROOT, "docs", "Cataclysm_SV定价表.xlsx")
PRESET_OUT = os.path.join(ROOT, "src", "main", "resources", "data", "meks", "sv", "cataclysm_preset.json")
PRESET = os.path.join(ROOT, "src", "main", "resources", "data", "meks", "sv", "emc_preset.json")

with open(os.path.join(DATA_DIR, "analysis.json"), encoding="utf-8") as f:
    ANALYSIS = json.load(f)
with open(PRESET, encoding="utf-8") as f:
    PRESET_EMC = {e["item"]: e["emc"] for e in json.load(f)}

# ---------------------------------------------------------------------------
# 显式定价表。sv=None 表示“暂不定价”，但可在这里写清原因。
# ---------------------------------------------------------------------------
PRICE = {
    # WIP / 未确认获取方式：sv 留空，只在逻辑里给占位建议
    "cataclysm:music_disc_the_cataclysmfarer": {"sv": None, "cat": "唱片", "logic": "当前版本未见掉落来源；若开放获取，建议 49152（6×钻石）。", "note": "不建议写入正式预设。"},
    "cataclysm:khopesh": {"sv": None, "cat": "武器", "logic": "WIP：未见配方/掉落；仅作为熔炼远古金属粒的原料出现在烧炼配方。", "note": "不建议写入正式预设。"},
    "cataclysm:coral_spear": {"sv": None, "cat": "武器", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:coral_bardiche": {"sv": None, "cat": "武器", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:belt_of_monstrosity": {"sv": None, "cat": "饰品", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:blazing_bone": {"sv": None, "cat": "材料", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:blood_clot": {"sv": None, "cat": "材料", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:urchin_spike": {"sv": None, "cat": "材料", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    "cataclysm:lionfish_spike": {"sv": None, "cat": "材料", "logic": "WIP：当前 3.32 未见配方/掉落。", "note": "不建议写入正式预设。"},
    # 常规材料
    "cataclysm:ancient_metal_nugget": {
        "sv": 455, "cat": "材料",
        "logic": "锚点：远古金属锭 4096 ÷ 9 ≈ 455；掉落自骸龙/沙漠遗迹。",
        "note": "与 9 粒合成锭的配方保持约等值，避免造币。",
    },
    "cataclysm:ancient_metal_ingot": {
        "sv": 4096, "cat": "材料",
        "logic": "锚点：16×铁锭(256)=4096，低于 1×钻石(8192)；骸龙系与沙漠遗迹中前期掉落。",
        "note": "黑钢、骸骨护甲的上游材料。",
    },
    "cataclysm:black_steel_nugget": {
        "sv": 910, "cat": "材料",
        "logic": "锚点：黑钢锭 8192 ÷ 9 ≈ 910；魂尸系掉落。",
        "note": "9 粒合成锭，等值即可。",
    },
    "cataclysm:black_steel_ingot": {
        "sv": 8192, "cat": "材料",
        "logic": "锚点：1×钻石(8192)；魂尸/遗迹掉落。",
        "note": "黑钢工具与咒魂线基础材料。",
    },
    "cataclysm:koboleton_bone": {
        "sv": 1024, "cat": "材料",
        "logic": "锚点：4×铁锭(256)；骸龙/骸龙斗士掉落。",
        "note": "骸骨护甲原料。",
    },
    "cataclysm:amethyst_crab_meat": {
        "sv": 256, "cat": "食物",
        "logic": "锚点：1×铁锭；紫水晶巨蟹掉落，带 5 秒再生。",
        "note": "仅作食物，价值不宜高于普通金苹果。",
    },
    "cataclysm:blessed_amethyst_crab_meat": {
        "sv": 256, "cat": "食物",
        "logic": "紫水晶祝福只消耗 120 秒，不消耗材料；建议与原料同价 256，避免反复祝福刷 SV。",
        "note": "如希望奖励祝福效果，可上调到 512，但要接受少量造币。",
    },
    "cataclysm:amethyst_crab_shell": {
        "sv": 2048, "cat": "材料",
        "logic": "锚点：8×铁锭(256)；紫水晶巨蟹掉落，护甲材料。",
        "note": "",
    },
    "cataclysm:chitin_claw": {
        "sv": 1024, "cat": "材料",
        "logic": "锚点：4×铁锭；巨钳守卫掉落。",
        "note": "",
    },
    "cataclysm:lacrima": {
        "sv": 4096, "cat": "材料",
        "logic": "锚点：16×铁锭；海洋结构宝箱与深海系生物掉落，海洋线核心货币。",
        "note": "泪宝石参与海石、风暴武器等大量配方。",
    },
    "cataclysm:essence_of_the_storm": {
        "sv": 16384, "cat": "材料",
        "logic": "锚点：2×钻石(8192)；斯库拉 BOSS 掉落。",
        "note": "风暴武器核心材料。",
    },
    "cataclysm:dying_ember": {
        "sv": 1024, "cat": "材料",
        "logic": "锚点：4×铁锭；炽燃狂魂掉落。",
        "note": "4 个可合成炽燃余烬。",
    },
    "cataclysm:burning_ashes": {
        "sv": 4096, "cat": "材料",
        "logic": "炽燃遗魂掉落；与 4×将熄余烬(1024) 的合成价一致。",
        "note": "",
    },
    "cataclysm:void_jaw": {
        "sv": 4096, "cat": "材料",
        "logic": "锚点：16×铁锭；末影甲虫掉落。",
        "note": "虚空散射箭原料。",
    },
    "cataclysm:void_core": {
        "sv": 32768, "cat": "材料",
        "logic": "锚点：4×钻石(8192)；末影傀儡 BOSS 掉落。",
        "note": "虚空融合线的关键掉落。",
    },
    "cataclysm:void_shard": {
        "sv": None, "cat": "材料",
        "logic": "WIP 占位：3.32 未见稳定掉落/配方；若开放获取，建议 8192（1×钻石）。",
        "note": "不建议写入正式预设。",
    },
    "cataclysm:chain_of_soul_binding": {
        "sv": None, "cat": "材料",
        "logic": "WIP 占位：未见稳定获取方式；若开放获取，建议 8192。",
        "note": "不建议写入正式预设。",
    },
    "cataclysm:crystallized_coral_fragments": {
        "sv": 512, "cat": "材料",
        "logic": "锚点：2×铁锭；珊瑚傀儡掉落。",
        "note": "4 个可合成晶化珊瑚。",
    },
    "cataclysm:coral_chunk": {
        "sv": 4096, "cat": "材料",
        "logic": "锚点：16×铁锭；珊瑚巨像掉落。",
        "note": "深渊祭品的替代材料。",
    },
    "cataclysm:athame": {
        "sv": 4096, "cat": "材料",
        "logic": "锚点：1×钻石；渊灵祭司/术士掉落。",
        "note": "深渊祭品必需材料。",
    },
    "cataclysm:strange_key": {
        "sv": 4096, "cat": "功能物品",
        "logic": "锚点：1×钻石；武弁掉落。",
        "note": "用于开启结构的钥匙，低值即可。",
    },
    "cataclysm:lionfish": {
        "sv": 128, "cat": "食物",
        "logic": "锚点：0.5×铁锭；蓑鲉掉落，毒+反胃+水下呼吸。",
        "note": "",
    },
    # 高阶级材料
    "cataclysm:witherite_ingot": {
        "sv": 24576, "cat": "材料",
        "logic": "先驱者掉落凋灵合金块，块价 221184 ÷ 9 = 24576；锚点 3×钻石。",
        "note": "科技线武器材料。",
    },
    "cataclysm:witherite_block": {
        "sv": 221184, "cat": "方块",
        "logic": "先驱者 BOSS 掉落；9×锭(24576) 推导值一致。",
        "note": "",
    },
    "cataclysm:enderite_ingot": {
        "sv": None, "cat": "材料",
        "logic": "WIP 占位：当前版本未见配方/掉落；若开放获取，建议 24576（3×钻石）。",
        "note": "不建议写入正式预设。",
    },
    "cataclysm:enderite_block": {
        "sv": None, "cat": "方块",
        "logic": "WIP 占位：与末影合金锭同规则；若开放获取，建议 221184（9×锭）。",
        "note": "不建议写入正式预设。",
    },
    "cataclysm:ignitium_ingot": {
        "sv": 49152, "cat": "材料",
        "logic": "焰魔 BOSS 掉落（3 次 roll）；锚点=6×钻石(8192)，与下界合金锭 50152 同级。",
        "note": "腾炎护甲/武器上游材料。",
    },
    "cataclysm:cursium_ingot": {
        "sv": 49152, "cat": "材料",
        "logic": "马勒迪克图斯 BOSS 掉落；与腾炎锭同档。",
        "note": "咒魂护甲/武器上游材料。",
    },
    "cataclysm:monstrous_horn": {
        "sv": 32768, "cat": "材料",
        "logic": "锚点：4×钻石；下界合金巨兽掉落。",
        "note": "恶兽头盔锻造材料。",
    },
    "cataclysm:lava_power_cell": {
        "sv": 16384, "cat": "材料",
        "logic": "锚点：2×钻石；下界合金巨兽掉落。",
        "note": "",
    },
    "cataclysm:netherite_effigy": {
        "sv": 64044, "cat": "材料",
        "logic": "合成材料和=64044（含下界合金锭 50152、不死图腾 9216 等）；同时是下界合金幼兽掉落，建议显式写同值。",
        "note": "显式值与合成价一致，不产生造币。",
    },
    # BOSS 唯一掉落
    "cataclysm:infernal_forge": {
        "sv": 262144, "cat": "武器",
        "logic": "BOSS 档：下界合金巨兽唯一掉落；锚点=32×钻石(8192)。",
        "note": "可通过机械融合砧继续合成高阶武器。",
    },
    "cataclysm:tidal_claws": {
        "sv": 262144, "cat": "武器",
        "logic": "BOSS 档：利维坦唯一掉落；锚点=32×钻石。",
        "note": "",
    },
    "cataclysm:gauntlet_of_guard": {
        "sv": 131072, "cat": "武器",
        "logic": "BOSS 档：末影守卫唯一掉落；锚点=16×钻石。",
        "note": "后续融合壁垒/沙暴护手。",
    },
    "cataclysm:sandstorm_in_a_bottle": {
        "sv": 131072, "cat": "功能物品",
        "logic": "BOSS 档：远古遗魂唯一掉落；锚点=16×钻石。",
        "note": "融合沙暴之怒/漩涡护手。",
    },
    "cataclysm:remnant_skull": {
        "sv": 65536, "cat": "材料",
        "logic": "BOSS 档：远古遗魂掉落；锚点=8×钻石。",
        "note": "",
    },
    "cataclysm:abyssal_egg": {
        "sv": 196608, "cat": "方块",
        "logic": "BOSS 档：利维坦/幼年利维坦掉落；锚点=24×钻石。",
        "note": "深渊系关键方块。",
    },
    # 唱片
    "cataclysm:music_disc_netherite_monstrosity": {"sv": 32768, "cat": "唱片", "logic": "下界合金巨兽 10% 掉落；锚点=4×钻石。", "note": ""},
    "cataclysm:music_disc_ender_guardian": {"sv": 32768, "cat": "唱片", "logic": "末影守卫掉落；同档 4×钻石。", "note": ""},
    "cataclysm:music_disc_ignis": {"sv": 49152, "cat": "唱片", "logic": "焰魔 10% 掉落；与腾炎锭同档 6×钻石。", "note": ""},
    "cataclysm:music_disc_the_harbinger": {"sv": 49152, "cat": "唱片", "logic": "先驱者掉落；同档 6×钻石。", "note": ""},
    "cataclysm:music_disc_the_leviathan": {"sv": 65536, "cat": "唱片", "logic": "利维坦 10% 掉落；BOSS 档 8×钻石。", "note": ""},
    "cataclysm:music_disc_ancient_remnant": {"sv": 49152, "cat": "唱片", "logic": "远古遗魂掉落；同档 6×钻石。", "note": ""},
    "cataclysm:music_disc_maledictus": {"sv": 49152, "cat": "唱片", "logic": "马勒迪克图斯掉落；同档 6×钻石。", "note": ""},
    "cataclysm:music_disc_scylla": {"sv": 49152, "cat": "唱片", "logic": "斯库拉掉落；同档 6×钻石。", "note": ""},
    # 锻造与融合产物
    "cataclysm:ignitium_helmet": {"sv": 156810, "cat": "护甲", "logic": "锻造：下界合金头盔 98609 + 腾炎锭 49152 + 模板 9049。", "note": "运行时不会自动推导锻造配方，需显式预设。"},
    "cataclysm:ignitium_chestplate": {"sv": 181386, "cat": "护甲", "logic": "锻造：下界合金胸甲 123185 + 腾炎锭 49152 + 模板 9049。", "note": ""},
    "cataclysm:ignitium_leggings": {"sv": 173194, "cat": "护甲", "logic": "锻造：下界合金护腿 114993 + 腾炎锭 49152 + 模板 9049。", "note": ""},
    "cataclysm:ignitium_boots": {"sv": 148618, "cat": "护甲", "logic": "锻造：下界合金靴子 90417 + 腾炎锭 49152 + 模板 9049。", "note": ""},
    "cataclysm:cursium_helmet": {"sv": 183930, "cat": "护甲", "logic": "锻造：下界合金头盔 98609 + 咒魂锭 49152 + 模板 36169。", "note": ""},
    "cataclysm:cursium_chestplate": {"sv": 208506, "cat": "护甲", "logic": "锻造：下界合金胸甲 123185 + 咒魂锭 49152 + 模板 36169。", "note": ""},
    "cataclysm:cursium_leggings": {"sv": 200314, "cat": "护甲", "logic": "锻造：下界合金护腿 114993 + 咒魂锭 49152 + 模板 36169。", "note": ""},
    "cataclysm:cursium_boots": {"sv": 175738, "cat": "护甲", "logic": "锻造：下界合金靴子 90417 + 咒魂锭 49152 + 模板 36169。", "note": ""},
    "cataclysm:monstrous_helm": {"sv": 138874, "cat": "护甲", "logic": "锻造：下界合金头盔 98609 + 恶兽犄角 32768 + 原版锻造模板 7497。", "note": ""},
    "cataclysm:ignitium_elytra_chestplate": {"sv": 279690, "cat": "护甲", "logic": "武器融合：腾炎胸甲 181386 + 鞘翅 98304（鞘翅建议补原版预设）。", "note": "原版 EMC 表没有鞘翅，需单独补 minecraft:elytra=98304。"},
    "cataclysm:brontes": {"sv": 307200, "cat": "武器", "logic": "武器融合：炼狱锻锤 262144 + 神怒长槊 45056。", "note": ""},
    "cataclysm:gauntlet_of_bulwark": {"sv": 232756, "cat": "武器", "logic": "武器融合：守卫者护手 131072 + 火焰壁垒 101684。", "note": ""},
    "cataclysm:gauntlet_of_maelstrom": {"sv": 262144, "cat": "武器", "logic": "武器融合：守卫者护手 131072 + 瓶中沙暴 131072。", "note": ""},
    "cataclysm:the_immolator": {"sv": 155648, "cat": "武器", "logic": "武器融合：歼灭战锤 106496 + 腾炎锭 49152。", "note": ""},
    "cataclysm:void_assault_shoulder_weapon": {"sv": 83264, "cat": "武器", "logic": "武器融合：凋灵突击肩炮 50496 + 虚空核心 32768。", "note": ""},
    "cataclysm:void_forge": {"sv": 294912, "cat": "武器", "logic": "武器融合：炼狱锻锤 262144 + 虚空核心 32768。", "note": ""},
    "cataclysm:wrath_of_the_desert": {"sv": 237604, "cat": "武器", "logic": "武器融合：咒魂弓 106532 + 瓶中沙暴 131072。", "note": ""},
}

# 明确不定价的原因（即使有掉落/配方，也先留空）
NO_PRICE_REASON = {
    "spawn_egg": "创造模式刷怪蛋，正常生存不可获取，不建议定价。",
    "bucket": "实体桶/生物桶，含 NBT 实体，SV 按 ID 记录会失真，不建议定价。",
    "creative_only": "结构/创造专属功能方块，不建议定价。",
    "wip": "当前版本未找到稳定配方或掉落，暂不定价。",
}

SOURCE_LABELS = {
    "chests": "箱子",
    "entities": "生物掉落",
    "blocks": "方块掉落",
    "archaeology": "考古",
}


def item_zh(item_id):
    name = item_id.split(":", 1)[1]
    for key in (f"item.cataclysm.{name}", f"block.cataclysm.{name}"):
        if key in ANALYSIS_LANG_ZH:
            return ANALYSIS_LANG_ZH[key]
    return item_zh_en(item_id)


def item_zh_en(item_id):
    name = item_id.split(":", 1)[1]
    for key in (f"item.cataclysm.{name}", f"block.cataclysm.{name}"):
        if key in ANALYSIS_LANG_EN:
            return ANALYSIS_LANG_EN[key]
    return name


def base_values():
    values = dict(PRESET_EMC)
    # ProjectE 表缺失的原版锚点
    values["minecraft:end_stone_brick_slab"] = 2
    values["minecraft:totem_of_undying"] = 9216
    values["minecraft:elytra"] = 98304
    values["minecraft:planks"] = 8  # 标签回退，取橡木木板
    for item_id, data in PRICE.items():
        if data["sv"]:
            values[item_id] = data["sv"]
    return values


def derive(values):
    recipes = [
        r for r in ANALYSIS["recipes"]
        if r["type"] in ("minecraft:crafting_shaped", "minecraft:crafting_shapeless")
    ]
    derived = {}
    for _ in range(80):
        changed = False
        for r in recipes:
            rid = r["result_id"]
            if rid in values:
                continue
            total = 0
            parts = []
            ok = True
            for ing in r["ingredients"]:
                key = ing.get("item") or ing.get("tag")
                v = values.get(key)
                if v is None:
                    ok = False
                    break
                total += v * ing.get("count", 1)
                parts.append((key, ing.get("count", 1), v))
            if not ok:
                continue
            per = total // max(1, r["result_count"])
            if per > 0:
                values[rid] = per
                derived[rid] = {
                    "total": total,
                    "count": r["result_count"],
                    "per": per,
                    "parts": parts,
                    "file": r["file"],
                }
                changed = True
        if not changed:
            break
    return derived


def formula_for_derived(info):
    parts = " + ".join(f"{c}×{k}({v})" for k, c, v in info["parts"])
    return f"合成：({parts}) ÷ {info['count']} = {info['per']}"


def source_text(item_id):
    name = item_id.split(":", 1)[1]
    bits = []
    recipe_types = set()
    for r in ANALYSIS["recipes"]:
        if r["result_id"] == item_id:
            recipe_types.add(r["type"])
    type_names = {
        "minecraft:crafting_shaped": "合成",
        "minecraft:crafting_shapeless": "合成",
        "minecraft:stonecutting": "切石",
        "minecraft:smithing_transform": "锻造",
        "cataclysm:weapon_fusion": "武器融合",
        "minecraft:smelting": "烧炼",
        "minecraft:blasting": "高炉",
        "cataclysm:amethyst_bless": "紫水晶祝福",
    }
    for t in sorted(recipe_types):
        bits.append(type_names.get(t, t))
    loot_bits = []
    for l in ANALYSIS["loot"]:
        if item_id in l["items"]:
            parts = l["file"].split("/")
            if len(parts) >= 2:
                kind = SOURCE_LABELS.get(parts[0], parts[0])
                loot_bits.append(f"{kind}:{parts[-1].removesuffix('.json')}")
    if loot_bits:
        bits.extend(loot_bits)
    return "；".join(bits) if bits else "未找到配方/掉落"


def classify(item):
    item_id = item["id"]
    if item_id in PRICE:
        return PRICE[item_id]["cat"]
    cls = item.get("class", "")
    name = item["name"]
    weapon_names = {
        "ancient_spear", "astrape", "bulwark_of_the_flame", "ceraunus", "cursed_bow",
        "laser_gatling", "meat_shredder", "soul_render", "the_annihilator", "the_incinerator",
        "wither_assault_shoulder_weapon", "void_assault_shoulder_weapon", "void_forge",
        "infernal_forge", "tidal_claws", "gauntlet_of_guard", "gauntlet_of_bulwark",
        "gauntlet_of_maelstrom", "brontes", "the_immolator", "wrath_of_the_desert",
        "coral_spear", "coral_bardiche", "khopesh", "the_incinerator",
    }
    if "SpawnEgg" in cls or name.endswith("spawn_egg"):
        return "刷怪蛋"
    if "Bucket" in cls or name.endswith("bucket"):
        return "桶"
    if cls == "RecordItem" or "music_disc" in name:
        return "唱片"
    if "Armor" in cls or name.endswith(("_helmet", "_chestplate", "_leggings", "_boots")) or "helm" in name or "pauldrons" in name:
        return "护甲"
    if name in weapon_names or "sword" in name or "spear" in name or "bow" in name or "gauntlet" in name or "forge" in name or "shredder" in name or "gatling" in name or "shoulder_weapon" in name or "halberd" in name or "hammer" in name or "claws" in name:
        return "武器"
    if "Sword" in cls or "Axe" in cls or "Pickaxe" in cls or "Shovel" in cls or "Hoe" in cls:
        return "工具/武器"
    if "Shield" in cls or "Targe" in cls:
        return "武器/盾"
    if name in ("aptrgangr_head", "draugr_head", "kobolediator_skull"):
        return "方块"
    if "Curios" in cls or name in ("belt_of_beginner", "belt_of_monstrosity", "blazing_grips", "sticky_gloves",
                                   "sturdy_boots", "necklace_of_the_desert", "vitality_ankh",
                                   "berserker_soul_amulet", "ring_of_grudged", "unbreakable_skull",
                                   "remnant_skull", "chain_of_soul_binding"):
        return "饰品"
    if name == "abyssal_sacrifice":
        return "功能物品"
    if name in ("crystallized_coral", "crystallized_coral_fragments", "coral_chunk"):
        return "材料"
    if "BlockItem" in cls or cls in ("BlockItemViaHelper", "CMBlockItem", "BlockItemCMRender"):
        return "方块"
    if "DungeonEyeItem" in cls or name.endswith("_eye"):
        return "进度道具"
    if "SmithingTemplate" in cls or name.endswith("smithing_template"):
        return "锻造模板"
    if "Food" in cls or name in ("lionfish", "amethyst_crab_meat", "blessed_amethyst_crab_meat"):
        return "食物"
    if name in ("void_scatter_arrow",):
        return "弹药"
    return "材料/其他"


def handling(item, derived):
    item_id = item["id"]
    if item_id in PRICE and PRICE[item_id]["sv"]:
        return "显式预设"
    if item_id in derived:
        return "自动推导"
    return "暂不定价"


def build_rows(values, derived):
    rows = []
    for item in ANALYSIS["items"]:
        item_id = item["id"]
        cat = classify(item)
        zh = item["zh"] or item_zh(item_id)
        en = item["en"] or item_zh_en(item_id)
        src = source_text(item_id)
        derived_info = derived.get(item_id)
        price = PRICE.get(item_id)
        treat = handling(item, derived)
        if price and price["sv"]:
            sv = price["sv"]
            logic = price["logic"]
            note = price.get("note", "")
        elif derived_info:
            sv = None
            logic = formula_for_derived(derived_info)
            note = "运行时按配方自动推导，无需写入预设表。"
        else:
            sv = None
            reason = ""
            note = ""
            if price:
                reason = price["logic"]
                note = price.get("note", "")
            elif cat == "刷怪蛋":
                reason = NO_PRICE_REASON["spawn_egg"]
            elif cat == "桶":
                reason = NO_PRICE_REASON["bucket"]
            elif item_id in {
                "cataclysm:dungeon_block", "cataclysm:altar_of_fire", "cataclysm:altar_of_void",
                "cataclysm:altar_of_amethyst", "cataclysm:altar_of_abyss", "cataclysm:boss_respawner",
                "cataclysm:cursed_tombstone", "cataclysm:emp", "cataclysm:door_of_seal",
                "cataclysm:goddess_statue", "cataclysm:door_of_seal_part",
            }:
                reason = NO_PRICE_REASON["creative_only"]
            else:
                reason = "暂无稳定获取方式或纯装饰/结构方块，暂不定价。"
            logic = reason
        rows.append({
            "cat": cat,
            "id": item_id,
            "zh": zh,
            "en": en,
            "source": src,
            "treat": treat,
            "sv": sv,
            "derived_sv": derived_info["per"] if derived_info else None,
            "logic": logic,
            "note": note,
        })
    return rows


def style_sheet(ws, headers, widths, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    values = base_values()
    derived = derive(values)
    rows = build_rows(values, derived)

    wb = Workbook()

    # 总览
    ws = wb.active
    ws.title = "定价原则"
    overview = [
        ["SV 定价原则", ""],
        ["数据基准", "L_Ender's Cataclysm 1.21.1-3.32；SV 采用 ProjectE EMC 风格，与 emc_preset.json 同一尺度。"],
        ["核心规则", "SV = 获取成本 + 稀有度 + 可重复性；可重复合成品的 SV 不得超过原料成本，否则会形成“合成造币”。"],
        ["自动推导", "原版合成(有型/无序)会被 MeksValues 自动按 Σ(原料SV×数量)÷输出数 推导，无需写入预设表。"],
        ["需显式预设", "锻造台、武器融合、紫水晶祝福等不会被运行时自动推导，表格会给出建议值；熔炉/高炉回收与切石机装饰多数暂不定价。"],
        ["暂不定价", "刷怪蛋、生物桶、创造专属功能方块、WIP/无稳定获取物品、纯装饰结构方块。"],
        ["", ""],
        ["价格锚点（ProjectE EMC）", ""],
        ["石头", "1"],
        ["铁锭", "256"],
        ["钻石", "8192"],
        ["下界合金锭", "50152"],
        ["下界合金胸甲", "123185"],
        ["下界之星", "139264"],
        ["", ""],
        ["Cataclysm 分档", ""],
        ["T0 常规材料", "1k-4k：骸龙骨、蟹壳、泪宝石、古代金属等"],
        ["T1 精英材料", "4k-16k：黑钢、珊瑚、风暴精华等"],
        ["T2 BOSS 材料", "24k-49k：凋灵合金、腾炎锭、咒魂锭等"],
        ["T3 BOSS 唯一掉落", "65k-307k：守卫者护手、炼狱锻锤、利维坦掉落等"],
        ["", ""],
        ["修改指引", ""],
        ["改基础材料", "编辑 design/cataclysm_sv/generate_cataclysm_sv_xlsx.py 的 PRICE 字典后重新运行。"],
        ["改合成产物", "一般不需要改；改上游材料 SV 后合成产物会自动重新推导。"],
        ["写入游戏", "把“建议SV”中需要显式预设的物品补进 emc_preset.json 或 Cataclysm 预设文件。"],
    ]
    for row in overview:
        ws.append(row)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    # 明细
    ws2 = wb.create_sheet("物品定价")
    headers = ["分类", "物品ID", "中文名", "英文名", "获取方式", "建议处理", "建议SV", "自动推导SV", "定价逻辑", "备注"]
    widths = [13, 38, 24, 30, 42, 14, 12, 13, 80, 40]
    ws2.append(headers)
    for r in sorted(rows, key=lambda x: (x["cat"], x["id"])):
        ws2.append([
            r["cat"], r["id"], r["zh"], r["en"], r["source"], r["treat"],
            r["sv"] if r["sv"] else "", r["derived_sv"] if r["derived_sv"] else "",
            r["logic"], r["note"],
        ])
    style_sheet(ws2, headers, widths)

    # 掉落/配方速查
    ws3 = wb.create_sheet("掉落与配方速查")
    headers3 = ["物品ID", "中文名", "掉落来源", "配方类型"]
    widths3 = [38, 24, 55, 40]
    ws3.append(headers3)
    recipes_by_id = defaultdict(list)
    for r in ANALYSIS["recipes"]:
        recipes_by_id[r["result_id"]].append(r["type"])
    loot_by_id = defaultdict(list)
    for l in ANALYSIS["loot"]:
        for it in l["items"]:
            loot_by_id[it].append(l["file"])
    for item_id in sorted(set(list(recipes_by_id) + list(loot_by_id))):
        zh = item_zh(item_id)
        loots = "；".join(loot_by_id.get(item_id, []))
        types = "；".join(sorted(set(recipes_by_id.get(item_id, []))))
        ws3.append([item_id, zh, loots, types])
    style_sheet(ws3, headers3, widths3)

    wb.save(OUT)
    preset = []
    seen = set()
    for r in rows:
        value = r["sv"] if r["sv"] else r["derived_sv"]
        if not value or r["id"] in seen:
            continue
        seen.add(r["id"])
        preset.append({"item": r["id"], "emc": value})
    preset.sort(key=lambda e: (e["item"].split(":")[0], e["item"]))
    with open(PRESET_OUT, "w", encoding="utf-8") as f:
        json.dump(preset, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print("rows", len(rows))
    print("preset entries", len(preset), "->", PRESET_OUT)
    print("explicit", sum(1 for r in rows if r["sv"]))
    print("derived", sum(1 for r in rows if r["derived_sv"] and not r["sv"]))
    print("unpriced", sum(1 for r in rows if not r["sv"] and not r["derived_sv"]))
    print("saved", OUT)


# 语言表直接内嵌，避免依赖临时目录
ANALYSIS_LANG_ZH = {}
ANALYSIS_LANG_EN = {}
for item in ANALYSIS["items"]:
    ANALYSIS_LANG_ZH[f"item.cataclysm.{item['name']}"] = item["zh"] or item["en"]
    ANALYSIS_LANG_EN[f"item.cataclysm.{item['name']}"] = item["en"] or item["zh"]
try:
    with open(os.path.join(ROOT, "src", "main", "resources", "assets", "meks", "lang", "zh_cn.json"), encoding="utf-8") as f:
        ANALYSIS_LANG_ZH.update(json.load(f))
except FileNotFoundError:
    pass
try:
    with open(os.path.join(ROOT, "src", "main", "resources", "assets", "meks", "lang", "en_us.json"), encoding="utf-8") as f:
        ANALYSIS_LANG_EN.update(json.load(f))
except FileNotFoundError:
    pass

if __name__ == "__main__":
    main()
